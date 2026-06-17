"""Read-only telemetry export helpers."""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Any


class TelemetryExportDatabaseError(RuntimeError):
    """Raised when telemetry export database reads fail."""


class TelemetryExportWorkbookError(RuntimeError):
    """Raised when telemetry export workbook generation fails."""


SESSION_COLUMNS = [
    "id",
    "app_session_id",
    "session_label",
    "started_at",
    "last_interaction_at",
    "closed_at",
    "status",
    "current_stage",
    "turns_count",
    "synthesis_generated",
    "pathways_generated",
    "pdf_downloaded",
    "problem_category",
    "engagement_signal",
    "feedback_submitted_at",
    "feedback_answer_1",
    "feedback_answer_2",
    "feedback_dropdown_values",
    "feedback_payload",
    "last_error",
    "created_at",
    "updated_at",
]

LLM_USAGE_COLUMNS = [
    "id",
    "app_session_id",
    "created_at",
    "llm_operation",
    "provider",
    "model",
    "input_tokens",
    "output_tokens",
    "total_tokens",
    "cached_input_tokens",
    "reasoning_tokens",
    "success",
    "latency_ms",
    "error_type",
    "error_message",
    "metadata",
]

SUMMARY_COLUMNS = [
    "session_id",
    "app_session_id",
    "session_label",
    "started_at",
    "last_interaction_at",
    "status",
    "current_stage",
    "turns_count",
    "synthesis_generated",
    "pathways_generated",
    "pdf_downloaded",
    "problem_category",
    "engagement_signal",
    "llm_calls",
    "input_tokens_total",
    "output_tokens_total",
    "total_tokens_total",
    "cached_input_tokens_total",
    "reasoning_tokens_total",
    "successful_llm_calls",
    "failed_llm_calls",
    "duration_seconds",
]


@dataclass(frozen=True)
class SheetData:
    name: str
    columns: list[str]
    rows: list[dict[str, Any]]


def build_telemetry_export_workbook(*, database_url: str, limit: int = 5_000) -> bytes:
    """Read telemetry rows from Postgres and return an Excel workbook as bytes."""
    bounded_limit = max(1, min(int(limit), 20_000))

    try:
        sheets = _fetch_export_data(database_url=database_url, limit=bounded_limit)
    except Exception as exc:
        raise TelemetryExportDatabaseError("telemetry_export_database_query_failed") from exc

    try:
        return _build_workbook_bytes(sheets)
    except Exception as exc:
        raise TelemetryExportWorkbookError("telemetry_export_workbook_generation_failed") from exc


def _fetch_export_data(*, database_url: str, limit: int) -> list[SheetData]:
    conn = None
    try:
        conn = _connect(database_url)
        with conn.cursor() as cursor:
            session_columns = _table_columns(cursor, "coach_sessions")
            usage_columns = _table_columns(cursor, "coach_llm_usage")

            sessions_sheet = _fetch_sessions_sheet(cursor, session_columns, limit)
            usage_sheet = _fetch_llm_usage_sheet(cursor, usage_columns, limit)
            summary_sheet = _fetch_session_token_summary_sheet(
                cursor,
                session_columns,
                usage_columns,
                limit,
            )

        return [sessions_sheet, usage_sheet, summary_sheet]
    finally:
        if conn is not None:
            conn.close()


def _connect(database_url: str) -> Any:
    try:
        import psycopg
    except Exception as exc:  # pragma: no cover - depends on environment
        raise RuntimeError("psycopg_unavailable") from exc

    return psycopg.connect(
        database_url,
        connect_timeout=5,
        options="-c statement_timeout=15000",
    )


def _table_columns(cursor: Any, table_name: str) -> set[str]:
    cursor.execute(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = current_schema()
          AND table_name = %s
        """,
        (table_name,),
    )
    return {str(row[0]) for row in cursor.fetchall()}


def _fetch_sessions_sheet(cursor: Any, available_columns: set[str], limit: int) -> SheetData:
    selected_columns = [column for column in SESSION_COLUMNS if column in available_columns]
    select_parts = [f"{column} AS {column}" for column in selected_columns]

    if {"started_at", "last_interaction_at"}.issubset(available_columns):
        end_column = "closed_at" if "closed_at" in available_columns else "last_interaction_at"
        select_parts.append(
            "EXTRACT(EPOCH FROM "
            f"(COALESCE({end_column}, last_interaction_at) - started_at)"
            ") AS duration_seconds"
        )
        selected_columns = [
            *selected_columns[: _duration_insert_index(selected_columns)],
            "duration_seconds",
            *selected_columns[_duration_insert_index(selected_columns) :],
        ]

    order_column = "started_at" if "started_at" in available_columns else "id"
    sql = f"""
        SELECT {", ".join(select_parts)}
        FROM coach_sessions
        ORDER BY {order_column} DESC
        LIMIT %s
    """
    cursor.execute(sql, (limit,))
    return SheetData("Sessions", selected_columns, _rows_from_cursor(cursor))


def _duration_insert_index(columns: list[str]) -> int:
    try:
        return columns.index("created_at")
    except ValueError:
        return len(columns)


def _fetch_llm_usage_sheet(cursor: Any, available_columns: set[str], limit: int) -> SheetData:
    selected_columns = [column for column in LLM_USAGE_COLUMNS if column in available_columns]
    select_parts = [f"{column} AS {column}" for column in selected_columns]
    order_column = "created_at" if "created_at" in available_columns else "id"

    sql = f"""
        SELECT {", ".join(select_parts)}
        FROM coach_llm_usage
        ORDER BY {order_column} DESC
        LIMIT %s
    """
    cursor.execute(sql, (limit,))
    return SheetData("LLM Usage", selected_columns, _rows_from_cursor(cursor))


def _fetch_session_token_summary_sheet(
    cursor: Any,
    session_columns: set[str],
    usage_columns: set[str],
    limit: int,
) -> SheetData:
    session_id_expr = _column_or_null("s", "id", session_columns, alias="session_id")
    session_parts = [
        session_id_expr,
        _column_or_null("s", "app_session_id", session_columns),
        _column_or_null("s", "session_label", session_columns),
        _column_or_null("s", "started_at", session_columns),
        _column_or_null("s", "last_interaction_at", session_columns),
        _column_or_null("s", "status", session_columns),
        _column_or_null("s", "current_stage", session_columns),
        _column_or_null("s", "turns_count", session_columns),
        _column_or_null("s", "synthesis_generated", session_columns),
        _column_or_null("s", "pathways_generated", session_columns),
        _column_or_null("s", "pdf_downloaded", session_columns),
        _column_or_null("s", "problem_category", session_columns),
        _column_or_null("s", "engagement_signal", session_columns),
    ]
    aggregate_parts = [
        "COUNT(u.id) AS llm_calls" if "id" in usage_columns else "0 AS llm_calls",
        _sum_or_zero("u", "input_tokens", usage_columns, "input_tokens_total"),
        _sum_or_zero("u", "output_tokens", usage_columns, "output_tokens_total"),
        _sum_or_zero("u", "total_tokens", usage_columns, "total_tokens_total"),
        _sum_or_zero("u", "cached_input_tokens", usage_columns, "cached_input_tokens_total"),
        _sum_or_zero("u", "reasoning_tokens", usage_columns, "reasoning_tokens_total"),
        (
            "COUNT(u.id) FILTER (WHERE u.success = TRUE) AS successful_llm_calls"
            if {"id", "success"}.issubset(usage_columns)
            else "0 AS successful_llm_calls"
        ),
        (
            "COUNT(u.id) FILTER (WHERE u.success = FALSE) AS failed_llm_calls"
            if {"id", "success"}.issubset(usage_columns)
            else "0 AS failed_llm_calls"
        ),
    ]
    duration_part = _duration_expression(session_columns)
    join_condition = (
        "u.app_session_id = s.app_session_id"
        if "app_session_id" in usage_columns and "app_session_id" in session_columns
        else "FALSE"
    )
    group_by_parts = [
        f"s.{column}"
        for column in (
            "id",
            "app_session_id",
            "session_label",
            "started_at",
            "last_interaction_at",
            "closed_at",
            "status",
            "current_stage",
            "turns_count",
            "synthesis_generated",
            "pathways_generated",
            "pdf_downloaded",
            "problem_category",
            "engagement_signal",
        )
        if column in session_columns
    ]
    group_by_clause = f"GROUP BY {', '.join(group_by_parts)}" if group_by_parts else ""
    order_column = "started_at" if "started_at" in session_columns else "id"

    sql = f"""
        WITH limited_sessions AS (
            SELECT *
            FROM coach_sessions
            ORDER BY {order_column} DESC
            LIMIT %s
        )
        SELECT
            {", ".join([*session_parts, *aggregate_parts, duration_part])}
        FROM limited_sessions s
        LEFT JOIN coach_llm_usage u
            ON {join_condition}
        {group_by_clause}
        ORDER BY {order_column} DESC
    """
    cursor.execute(sql, (limit,))
    return SheetData("Session Token Summary", SUMMARY_COLUMNS, _rows_from_cursor(cursor))


def _column_or_null(
    table_alias: str,
    column: str,
    available_columns: set[str],
    *,
    alias: str | None = None,
) -> str:
    output_name = alias or column
    if column in available_columns:
        return f"{table_alias}.{column} AS {output_name}"

    return f"NULL AS {output_name}"


def _sum_or_zero(
    table_alias: str,
    column: str,
    available_columns: set[str],
    alias: str,
) -> str:
    if column in available_columns:
        return f"COALESCE(SUM({table_alias}.{column}), 0) AS {alias}"

    return f"0 AS {alias}"


def _duration_expression(session_columns: set[str]) -> str:
    if {"started_at", "last_interaction_at"}.issubset(session_columns):
        end_column = "closed_at" if "closed_at" in session_columns else "last_interaction_at"
        return (
            "EXTRACT(EPOCH FROM "
            f"(COALESCE(s.{end_column}, s.last_interaction_at) - s.started_at)"
            ") AS duration_seconds"
        )

    return "NULL AS duration_seconds"


def _rows_from_cursor(cursor: Any) -> list[dict[str, Any]]:
    column_names = [description[0] for description in cursor.description or []]
    return [dict(zip(column_names, row, strict=False)) for row in cursor.fetchall()]


def _build_workbook_bytes(sheets: list[SheetData]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_data in sheets:
        sheet = workbook.create_sheet(title=sheet_data.name)
        sheet.append(sheet_data.columns)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"

        for row in sheet_data.rows:
            sheet.append([_excel_safe_value(row.get(column)) for column in sheet_data.columns])

        if sheet.max_column > 0:
            sheet.auto_filter.ref = sheet.dimensions
            _apply_column_widths(sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _excel_safe_value(value: Any) -> Any:
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)

    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, dict | list | tuple):
        return json.dumps(value, default=str, sort_keys=True)

    return value


def _apply_column_widths(sheet: Any) -> None:
    for column_cells in sheet.columns:
        header = str(column_cells[0].value or "")
        max_length = len(header)
        for cell in column_cells[1:]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 60))
        sheet.column_dimensions[column_cells[0].column_letter].width = min(
            max(max_length + 2, 10),
            64,
        )
