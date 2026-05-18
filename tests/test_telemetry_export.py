import unittest
from datetime import datetime, timezone
from io import BytesIO
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.api import app
from backend.telemetry.export import SheetData, _build_workbook_bytes


class TelemetryExportRouteTests(unittest.TestCase):
    def setUp(self) -> None:
        self.client = TestClient(app)

    @patch.dict(
        "os.environ",
        {
            "TELEMETRY_EXPORT_TOKEN": "test-export-token",
            "TELEMETRY_DATABASE_URL": "postgresql://example",
        },
        clear=False,
    )
    @patch("backend.admin.telemetry_export_routes.build_telemetry_export_workbook")
    def test_export_route_returns_xlsx_download(self, mock_build_workbook) -> None:
        mock_build_workbook.return_value = b"xlsx-bytes"

        response = self.client.get(
            "/admin/telemetry/export.xlsx?token=test-export-token&limit=1000"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("aether-glimpse-telemetry-", response.headers["content-disposition"])
        self.assertEqual(response.content, b"xlsx-bytes")
        mock_build_workbook.assert_called_once_with(
            database_url="postgresql://example",
            limit=1000,
        )

    @patch.dict(
        "os.environ",
        {
            "TELEMETRY_EXPORT_TOKEN": "test-export-token",
            "TELEMETRY_DATABASE_URL": "postgresql://example",
        },
        clear=False,
    )
    def test_export_route_rejects_wrong_token(self) -> None:
        response = self.client.get("/admin/telemetry/export.xlsx?token=wrong")

        self.assertEqual(response.status_code, 403)

    @patch.dict("os.environ", {"TELEMETRY_DATABASE_URL": "postgresql://example"}, clear=True)
    def test_export_route_requires_configured_token(self) -> None:
        response = self.client.get("/admin/telemetry/export.xlsx?token=test-export-token")

        self.assertEqual(response.status_code, 503)


class TelemetryExportWorkbookTests(unittest.TestCase):
    def test_workbook_contains_required_sheets_and_readable_values(self) -> None:
        workbook_bytes = _build_workbook_bytes(
            [
                SheetData(
                    name="Sessions",
                    columns=["id", "app_session_id", "created_at", "feedback_payload"],
                    rows=[
                        {
                            "id": 1,
                            "app_session_id": "session-1",
                            "created_at": datetime(2026, 5, 18, tzinfo=timezone.utc),
                            "feedback_payload": {"safe": True},
                        }
                    ],
                ),
                SheetData(
                    name="LLM Usage",
                    columns=["id", "app_session_id", "total_tokens"],
                    rows=[{"id": 7, "app_session_id": "session-1", "total_tokens": 42}],
                ),
                SheetData(
                    name="Session Token Summary",
                    columns=["session_id", "app_session_id", "total_tokens_total"],
                    rows=[
                        {
                            "session_id": 1,
                            "app_session_id": "session-1",
                            "total_tokens_total": 42,
                        }
                    ],
                ),
            ]
        )

        workbook = load_workbook(BytesIO(workbook_bytes))

        self.assertEqual(
            workbook.sheetnames,
            ["Sessions", "LLM Usage", "Session Token Summary"],
        )
        self.assertEqual(workbook["Sessions"]["A1"].value, "id")
        self.assertEqual(workbook["Sessions"]["D2"].value, '{"safe": true}')
        self.assertEqual(workbook["Sessions"]["A1"].font.bold, True)
        self.assertEqual(workbook["LLM Usage"].freeze_panes, "A2")


if __name__ == "__main__":
    unittest.main()
